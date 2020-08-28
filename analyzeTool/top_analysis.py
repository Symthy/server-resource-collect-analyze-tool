"""
analyze top log and output csv file (row: time, column: process Id and command name)
    option:
        --withExcel : Output Excel File(include line graph) and csv file
        --viewGraph : View line graph (memory use rate and cpu use rate)
        --startTime "YYYY-mm-dd" : output start date time filter
        --endTime "YYYY-mm-dd" : output end date time filter
"""

import glob
import re
import sys
import csv
import openpyxl
import matplotlib.pyplot as plt
from typing import List, Dict, Optional
from openpyxl.chart import LineChart, Reference, Series
import datetime as dt

# Fixed value
PID_INDEX = 0
CPU_INDEX = 8
MEM_INDEX = 9
COMMAND_INDEX = 11
PID_VALUE_COLUMN_COUNT = 12
EXCEL_OPTION = '--withExcel'
VIEW_GRAPH_OPTION = '--viewGraph'
START_DATETIME_OPTION = '--startTime'
END_DATETIME_OPTION = '--endTime'

# Changeable values
FILTER_VALUE = 1.0  # Output only values more than this value
RANK_TOP_LIMIT = 5  # Maximum number of processes to output
OUTPUT_TOP_MEM_NAME = 'top_memory_result'  # Output file name and graph title
OUTPUT_TOP_CPU_NAME = 'top_cpu_result'  # Output file name and graph title


def view_line_graph(name: str, header: List[str], big_order_indexes: List[int], array2d: List[List[str]]):
    """
    Description:
        create and view line graph by matplotlib.
    :param name: graph name.
    :param header: top line of output file. top line is '' and process id.
    :param big_order_indexes: column(process id) index list in big order of max value per process id.
    :param array2d: 2d array (row: date time, column: process). row 0 is date time.
    :return: void
    """
    times = [array2d[i][0] for i in range(len(array2d)) if i < len(array2d) - 1]
    x_alias = [i for i in times[::int(len(times) / 10)]]
    fig, axes = plt.subplots()
    for col in range(len(big_order_indexes)):
        if col == 0:
            # skip because column 0 is time
            continue
        if col > RANK_TOP_LIMIT or col > len(big_order_indexes):
            break
        y_values = []
        for row in range(len(array2d)):
            if not row < len(array2d) - 1:
                break
            y_values.append(
                float(array2d[row][big_order_indexes[col]]) if array2d[row][big_order_indexes[col]] != '' else None)
        axes.plot(times, y_values, label=header[col])
    axes.set_title(name)
    axes.set_xlabel('Time')
    axes.set_xticks(x_alias)
    axes.set_ylabel('Use Rate [%]')
    axes.set_ylim(0, 100)
    axes.legend()
    axes.grid()
    plt.xticks(rotation=30)


def create_excel_line_graph(sheet, rows_num: int):
    """
    Description:
        Undone Function (Because can't check the operation in the developer environment).
        create line graph in excel file.
    :param sheet: excel sheet object.
    :param rows_num: row count.
    :return: void
    """
    chart = openpyxl.chart.LineChart()
    chart.title = 'プロセス毎の使用率'
    chart.x_axis.title = "時刻"
    chart.y_axis.title = '使用量 [%]'
    chart.height = 16
    chart.width = 24
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    time_refs = Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=rows_num - 1)
    pid_refs = Reference(sheet, min_col=2, max_col=RANK_TOP_LIMIT, min_row=1, max_row=rows_num - 1)
    chart.add_data(pid_refs, titles_from_data=True)
    chart.set_categories(time_refs)
    sheet.add_chart(chart, "A" + str(rows_num + 3))


def write_excel_file(filename, header, big_order_indexes, array2d):
    """
    Description:
        output excel file (row: date time, column: process)
    :param filename: output file name.
    :param header: top line of output file. top line is '' and process id.
    :param big_order_indexes: column(process id) index list in big order of max value per process id.
    :param array2d: 2d array (row: date time, column: process). row 0 is date time.
    :return: void
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'result'
    sheet = wb['result']
    for i in range(len(header)):
        sheet.cell(row=1, column=i + 1).value = header[i]
    for row in range(len(array2d)):
        for col in range(len(big_order_indexes)):
            sheet.cell(row=row + 2, column=col + 1).value = array2d[row][big_order_indexes[col]]
    create_excel_line_graph(sheet, len(array2d))
    wb.save('../output/' + filename + '.xlsx')


def write_csv_file(filename: str, header: List[str], big_order_indexes: List[int], array2d: List[List[str]]):
    """
    Description:
        output csv file (row: date time, column: process)
    :param filename: output file name.
    :param header: top line of output file. top line is '' and process id.
    :param big_order_indexes: column(process id) index list in big order of max value per process id.
    :param array2d: 2d array (row: date time, column: process). row 0 is date time.
    :return: void
    """
    with open('../output/' + filename + '.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile, lineterminator='\n')
        writer.writerow(header)
        # writer.writerows(array2d)
        for row in range(len(array2d)):
            writer.writerow([array2d[row][index] for index in big_order_indexes])


def create_max_value_order(max_values_per_pid: List[str]) -> List[int]:
    """
    Description:
        create column(process id) index list in big order of max value.
        index 0 is date time column.
    :param max_values_per_pid: max value list per process id
    :return: column(process id) index list in big order of max value per process id
    """
    max_values: List[float] = list(map(lambda s: float(s), max_values_per_pid))
    sorted_max_values = sorted(max_values, reverse=True)
    big_order_indexes: List[int] = [max_values.index(v) + 1 for v in sorted_max_values]
    return [0] + big_order_indexes


def create_max_value_per_pid(array2d: List[List[str]]) -> List[str]:
    """
    Description:
        create max value list per process id.
    :param array2d: 2d array (row: date time, column: process). row 0 is date time.
    :return: max value list per process id.
    """
    rows_len: int = len(array2d)
    columns_len: int = len(array2d[0])
    max_values: List[str] = []
    for col in range(columns_len):
        if col == 0:
            continue
        max = '0.0'
        for row in range(rows_len):
            if array2d[row][col] != '' and float(max) < float(array2d[row][col]):
                max = array2d[row][col]
        max_values.append(max)
    return max_values


def fill_empty_string(pid_count: int, array2d: List[List[str]]):
    """
    Description:
        fill empty string on each row of array2d. the number to fill is the number of missing columns.
    :param pid_count: process id count.
    :param array2d: 2d array (row: date time, column: process). 0 row is date time.
    :return: void
    """
    for record in array2d:
        diff = pid_count - len(record)
        record.extend([''] * diff)


def write_file_and_view_graph(name: str, pids_header: List[str], array2d: List[List[str]], is_output_excel: bool,
                              is_view_graph: bool):
    """
    Description:
        output file and view graph.
    :param name: file name and graph name
    :param pids_header: list of process id. index 0 is ''.
    :param array2d: 2d array (row: date time, column: process). 0 row is date time.
    :param is_output_excel: output excel file flag.
    :param is_view_graph: view line graph flag.
    :return: void
    """
    max_values_per_pid: List[str] = create_max_value_per_pid(array2d)
    big_order_indexes: List[int] = create_max_value_order(max_values_per_pid)
    array2d.append(['MAX:'] + max_values_per_pid)
    write_csv_file(name, pids_header, big_order_indexes, array2d)
    if is_output_excel:
        write_excel_file(name, pids_header, big_order_indexes, array2d)
    if is_view_graph:
        view_line_graph(name, pids_header, big_order_indexes, array2d)


def except_out_of_start_to_end_filter_range(filter_start_time: Optional, filter_end_time: Optional,
                                            pids_header: List[str], array2d: List[List[str]]):
    """
    Description:
        except out of start date time to end date time range
    :param filter_start_time: output to start date time.
    :param filter_end_time: output to end date time.
    :param pids_header: process id list. first is ''.
    :param array2d: 0 row is date time.
    :return: void
    """
    if filter_start_time is None and filter_end_time is None:
        return
    # remove out of start to end range
    for row in reversed(range(len(array2d) - 1)):
        date_time: dt = dt.datetime.strptime(array2d[row][0], '%Y-%m-%d %H:%M:%S')
        if filter_start_time is not None and date_time < filter_start_time:
            del (array2d[row])
        if filter_end_time is not None and date_time > filter_end_time:
            del (array2d[row])
    # remove all empty column
    delete_column_indexes = []
    for col in reversed(range(len(array2d[0]) - 1)):
        is_empty_column = True
        for row in range(len(array2d)):
            if array2d[row][col] != '':
                is_empty_column = False
                break
        if is_empty_column:
            delete_column_indexes.append(col)
    for col in delete_column_indexes:
        del (pids_header[col])
        for row in reversed(range(len(array2d) - 1)):
            del (array2d[row][col])


def add_time_and_value_array2d(datetime: str, pids: List[str], value_dict: Dict, array2d: List[List[str]]):
    """
    Discription:
        add date time and value to 2d array.
    :param datetime: current date and time of log one line.
    :param pids: list of process id.
    :param value_dict: map of process id and value.
    :param array2d: 2d array (row: date time, column: process). 0 row is date time.
    :return: void
    """
    if datetime == '':
        return
    record: List[str] = [''] * (len(pids) + 1)
    record[0] = datetime
    for pid_key in value_dict.keys():
        record[pids.index(pid_key) + 1] = value_dict[pid_key]
    array2d.append(record)  # array2d record : datetime + pids


def add_time_and_mem_cpu_array2d(datetime: str, mem_pids: List[str], cpu_pids: List[str], mem_dict: Dict,
                                 cpu_dict: Dict, time_mem_array2d: List[List[str]], time_cpu_array2d: List[List[str]]):
    """
    Description:
        add date time and value (memory use rate or cpu use rate) to 2d array.
    :param datetime: current date and time of log one line.
    :param mem_pids: list of process id (for memory use rate).
    :param cpu_pids: list of process id (for cpu use rate).
    :param mem_dict: map of process id and memory use rate.
    :param cpu_dict:  map of process id and cpu use rate.
    :param time_mem_array2d: 2d array of memory use rate (row: date time, column: process). 0 row is date time.
    :param time_cpu_array2d: 2d array of cpu use rate (row: date time, column: process). 0 row is date time.
    :return: void
    """
    add_time_and_value_array2d(datetime, mem_pids, mem_dict, time_mem_array2d)
    add_time_and_value_array2d(datetime, cpu_pids, cpu_dict, time_cpu_array2d)


def analyze_pid_value_line(line: str, pids: List[str], value_dict: Dict, target_index: int):
    """
    Description:
        analyze one line of process information.
    :param line: log line.
    :param pids: list of process id.
    :param value_dict: map of process id and value.
    :param target_index: index of target value in log line.
    :return: void
    """
    line_columns: List[str] = line.split()
    pid: str = line_columns[PID_INDEX] + '(' + line_columns[COMMAND_INDEX] + ')'
    value: str = line_columns[target_index]
    if float(value) >= FILTER_VALUE:
        if pid not in pids:
            pids.append(pid)
        value_dict[pid] = value


current_date: str = ''  # format is 'YYYY-mm-dd'
is_zero_hour: bool = False


def get_current_date_time(start_date: str, current_time: str) -> str:
    """
    Description:
        create current date time of target log line
    :param start_date: log start date. format is 'YYYYmmdd'
    :param current_time: current time of target log line. format is 'HH:MM:ss'
    :return: current date and time
    """
    global current_date, is_zero_hour
    if current_date == '':
        current_date = dt.datetime.strptime(start_date, '%Y%m%d').strftime('%Y-%m-%d')
        is_zero_hour = current_time.startswith('00:')
    if is_zero_hour and current_time.startswith('01:'):
        is_zero_hour = False
    if not is_zero_hour and current_time.startswith('00:'):
        is_zero_hour = True
        dt_current_date = dt.datetime.strptime(current_date, '%Y-%m-%d') + dt.timedelta(days=1)
        current_date = dt_current_date.strftime('%Y-%m-%d')
    return current_date + ' ' + current_time


def analyze_top_log_lines(start_date: str, lines: List[str], mem_pids: List[str], cpu_pids: List[str],
                          time_mem_array2d: List[List[str]], time_cpu_array2d: List[List[str]]):
    """
    Description:
        analyze top log lines. create 2d array memory use rate and cpu use rate.
    :param start_date: log start date.
    :param lines: log file lines.
    :param mem_pids: list of memory use rate per process.
    :param cpu_pids: list of cpu use rate per process.
    :param time_mem_array2d: 2d array (row: date time, column: process). 0 row is date time.
    :param time_cpu_array2d: 2d array (row: date time, column: process). 0 row is date time.
    :return: void
    """
    is_pid_value_block = False
    mem_dict: Dict = {}
    cpu_dict: Dict = {}
    time = ''
    for line in lines:
        line_columns = line.split();
        if line.startswith('top -'):
            datetime = get_current_date_time(start_date, line_columns[2])
            add_time_and_mem_cpu_array2d(datetime, mem_pids, cpu_pids, mem_dict, cpu_dict, time_mem_array2d,
                                         time_cpu_array2d)
            is_pid_value_block = False
            mem_dict = {}
            cpu_dict = {}
            continue
        if line.startswith('    PID'):
            is_pid_value_block = True
            continue
        if is_pid_value_block and len(line_columns) == PID_VALUE_COLUMN_COUNT:
            analyze_pid_value_line(line, mem_pids, mem_dict, MEM_INDEX)
            analyze_pid_value_line(line, cpu_pids, cpu_dict, CPU_INDEX)
    add_time_and_mem_cpu_array2d(time, mem_pids, cpu_pids, mem_dict, cpu_dict, time_mem_array2d, time_cpu_array2d)


def analyze_top_log(file_path: str, is_output_excel: bool, is_view_graph: bool, filter_start_time: dt,
                    filter_end_time: dt):
    """

    :param file_path:
    :param is_output_excel:
    :param is_view_graph:
    :param filter_end_time:
    :param filter_start_time:
    :return: void
    """
    mem_pids: List[str] = []
    cpu_pids: List[str] = []
    time_mem_array2d: List[List[str]] = []
    time_cpu_array2d: List[List[str]] = []
    start_date = re.search(r'\d+', file_path).group()
    with open(file_path, 'r') as f:
        lines = f.readlines()
        analyze_top_log_lines(start_date, lines, mem_pids, cpu_pids, time_mem_array2d, time_cpu_array2d)
    mem_pids = [''] + mem_pids
    cpu_pids = [''] + cpu_pids
    fill_empty_string(len(mem_pids), time_mem_array2d)
    fill_empty_string(len(cpu_pids), time_cpu_array2d)
    except_out_of_start_to_end_filter_range(filter_start_time, filter_end_time, mem_pids, time_mem_array2d)
    except_out_of_start_to_end_filter_range(filter_start_time, filter_end_time, cpu_pids, time_cpu_array2d)
    write_file_and_view_graph(OUTPUT_TOP_MEM_NAME, mem_pids, time_mem_array2d, is_output_excel, is_view_graph)
    write_file_and_view_graph(OUTPUT_TOP_CPU_NAME, cpu_pids, time_cpu_array2d, is_output_excel, is_view_graph)
    if is_view_graph:
        plt.show()


def convert_date_time(option: str, args: List[str]):
    """

    :param option: option name
    :param args: command line arguments
    :return: void
    """
    if not option in args:
        return None
    index = args.index(option)
    if len(args) > index:
        filter_start_time = args[index + 1]
        try:
            return dt.datetime.strptime(filter_start_time, '%Y-%m-%d %H:%M:%S')
        except Exception:
            raise


def main(args: List[str]):
    """

    :param args: command line arguments
    :return: void
    """
    is_output_excel = False
    is_view_graph = False
    filter_start_time: Optional = None
    filter_end_time: Optional = None
    if EXCEL_OPTION in args:
        is_output_excel = True
    if VIEW_GRAPH_OPTION in args:
        is_view_graph = True
    if START_DATETIME_OPTION in args:
        try:
            filter_start_time = convert_date_time(START_DATETIME_OPTION, args)
        except Exception:
            print('invalid --startTime format (YYYY-mm-dd HH:MM:SS)')
            raise
    if END_DATETIME_OPTION in args:
        try:
            filter_end_time = convert_date_time(END_DATETIME_OPTION, args)
        except Exception:
            print('invalid --endTime format (YYYY-mm-dd HH:MM:SS)')
            raise
    file_paths: List[str] = glob.glob("../input/top_*.log")
    for file_path in file_paths:
        analyze_top_log(file_path, is_output_excel, is_view_graph, filter_start_time, filter_end_time)


main(sys.argv)
